import openpyxl
import json
import sys
import os
import re

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_FILE = 'Coto JLPT Grammar and Vocab Mega List.xlsx'
VOCAB_OUT = 'src/data/vocab_database.json'
GRAMMAR_OUT = 'src/data/grammar_database.json'

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Helper function to generate standard Japanese conjugation forms for verbs/adjectives/words
def generate_conjugations(kanji, reading, meaning):
    text = kanji if kanji else reading
    r_text = reading if reading else kanji

    # Simple heuristic to determine POS:
    # 1. Ends in i (not noun): i-adj
    # 2. Ends in u, ku, su, tsu, nu, bu, mu, ru, gu: verb
    # 3. Suru verb: ends in する
    # Default fallback returns standard form mappings
    
    root = text
    root_r = r_text
    
    # 1. Suru verb e.g. 勉強する -> べんきょうする
    if text.endswith('する') or r_text.endswith('する'):
        stem = text[:-2]
        stem_r = r_text[:-2]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'した',
            'pastPosReading': stem_r + 'した',
            'pastNeg': stem + 'しなかった',
            'pastNegReading': stem_r + 'しなかった',
            'teForm': stem + 'して',
            'teFormReading': stem_r + 'して',
            'taiForm': stem + 'したい',
            'taiFormReading': stem_r + 'したい',
        }
    
    # 2. Ichidan / Godan verbs ending in ru/ku/su/etc.
    # Verb rules for quick standard conjugation:
    if text.endswith('い') and not (text.endswith('ない') or text.endswith('きれい')): # i-adjective
        stem = text[:-1]
        stem_r = r_text[:-1]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'かった',
            'pastPosReading': stem_r + 'かった',
            'pastNeg': stem + 'くなかった',
            'pastNegReading': stem_r + 'くなかった',
            'teForm': stem + 'くて',
            'teFormReading': stem_r + 'くて',
            'taiForm': text + ' (i-adj)',
            'taiFormReading': r_text + ' (i-adj)',
        }
    
    if text.endswith('る') or r_text.endswith('る'):
        stem = text[:-1]
        stem_r = r_text[:-1]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'た',
            'pastPosReading': stem_r + 'た',
            'pastNeg': stem + 'ない',
            'pastNegReading': stem_r + 'ない',
            'teForm': stem + 'て',
            'teFormReading': stem_r + 'て',
            'taiForm': stem + 'たい',
            'taiFormReading': stem_r + 'たい',
        }
    
    if text.endswith('く') or r_text.endswith('く'):
        stem = text[:-1]
        stem_r = r_text[:-1]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'いた',
            'pastPosReading': stem_r + 'いた',
            'pastNeg': stem + 'かない',
            'pastNegReading': stem_r + 'かない',
            'teForm': stem + 'いて',
            'teFormReading': stem_r + 'いて',
            'taiForm': stem + 'きたい',
            'taiFormReading': stem_r + 'きたい',
        }
    
    if text.endswith('む') or text.endswith('ぶ') or text.endswith('ぬ'):
        stem = text[:-1]
        stem_r = r_text[:-1]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'んだ',
            'pastPosReading': stem_r + 'んだ',
            'pastNeg': stem + 'まない',
            'pastNegReading': stem_r + 'まない',
            'teForm': stem + 'んで',
            'teFormReading': stem_r + 'んで',
            'taiForm': stem + 'みたい',
            'taiFormReading': stem_r + 'みたい',
        }
    
    if text.endswith('す') or r_text.endswith('す'):
        stem = text[:-1]
        stem_r = r_text[:-1]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'した',
            'pastPosReading': stem_r + 'した',
            'pastNeg': stem + 'さない',
            'pastNegReading': stem_r + 'さない',
            'teForm': stem + 'して',
            'teFormReading': stem_r + 'して',
            'taiForm': stem + 'したい',
            'taiFormReading': stem_r + 'したい',
        }
    
    if text.endswith('つ') or r_text.endswith('つ') or text.endswith('う') or r_text.endswith('う'):
        stem = text[:-1]
        stem_r = r_text[:-1]
        return {
            'root': text,
            'rootReading': r_text,
            'short': text,
            'shortReading': r_text,
            'pastPos': stem + 'った',
            'pastPosReading': stem_r + 'った',
            'pastNeg': stem + 'わない',
            'pastNegReading': stem_r + 'わない',
            'teForm': stem + 'って',
            'teFormReading': stem_r + 'って',
            'taiForm': stem + 'いたい',
            'taiFormReading': stem_r + 'いたい',
        }

    # Noun / general word default copula forms
    return {
        'root': text,
        'rootReading': r_text,
        'short': text + 'だ',
        'shortReading': r_text + 'だ',
        'pastPos': text + 'だった',
        'pastPosReading': r_text + 'だった',
        'pastNeg': text + 'ではない',
        'pastNegReading': r_text + 'ではない',
        'teForm': text + 'で',
        'teFormReading': r_text + 'で',
        'taiForm': text + 'になりたい',
        'taiFormReading': r_text + 'になりたい',
    }

# Process Vocab
vocab_list = []
for sheet_name in wb.sheetnames:
    if 'Vocab' not in sheet_name:
        continue
    level = sheet_name.split()[0]
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    
    for i, r in enumerate(rows[2:]):
        if not r or (not r[0] and not r[1]) or not r[2]:
            continue
        kanji = str(r[0]).strip() if r[0] is not None else ''
        reading = str(r[1]).strip() if r[1] is not None else ''
        meaning = str(r[2]).strip() if r[2] is not None else ''
        
        if not reading and not kanji:
            continue
            
        conjs = generate_conjugations(kanji, reading, meaning)
        
        vocab_list.append({
            'id': f'v_{level.lower()}_{i}',
            'kanji': kanji,
            'reading': reading or kanji,
            'meaning': meaning,
            'jlpt': level,
            'type': 'vocab',
            'conjugations': conjs
        })

print(f"Extracted {len(vocab_list)} vocab entries.")

# Helper to generate Furigana-Only Sample Sentences and Nuance Explanations for Grammar Points
def enrich_grammar_point(title, level):
    # Sanitize title
    clean_title = title.replace('～', '').replace('~', '').strip()
    
    # Structure default
    structure = f"Verb [Plain Form] + {clean_title} / Noun + {clean_title}"
    
    # Nuance explanation
    nuance = f"Used in {level} Japanese to express concepts related to '{clean_title}'. It clarifies connections, conditionals, intentions, or state transitions in sentences."
    
    # Specific curated nuances & furigana-only sentences for popular patterns
    if 'から' in title:
        structure = "Verb / Noun / Adjective + から"
        nuance = "Expresses cause or reason ('because', 'since'). Gives a subjective, personal reason."
        samples = [
            {"furigana": "あめ が ふっている から、 うち に います。", "english": "Because it is raining, I am staying at home."},
            {"furigana": "おなかが すいた から、 ごはん を たべます。", "english": "Since I am hungry, I will eat a meal."}
        ]
    elif 'てはいけない' in title or 'てはならない' in title:
        structure = "Verb [Te-form] + はいいけない"
        nuance = "Expresses strong prohibition or forbidding an action ('must not', 'should not')."
        samples = [
            {"furigana": "ここ で たばこ を すって は いけません。", "english": "You must not smoke cigarettes here."},
            {"furigana": "テスト の とき に はなして は いけません。", "english": "You must not talk during the test."}
        ]
    elif 'たび' in title:
        structure = "Verb [Dictionary form] + たび / Noun + の + たび"
        nuance = "Expresses 'every time' or 'whenever' something happens, a specific result always follows."
        samples = [
            {"furigana": "りょこう に いく たび に、 おみやげ を かいます。", "english": "Whenever I go on a trip, I buy souvenirs."},
            {"furigana": "この しゃしん を みる たび に、 こども の ころ を おもいだします。", "english": "Every time I look at this photo, I remember my childhood."}
        ]
    elif 'ようにする' in title:
        structure = "Verb [Dictionary/Nai form] + ようにする"
        nuance = "Expresses making an effort or trying one's best to establish a habit ('make sure to', 'try to')."
        samples = [
            {"furigana": "まいにち やさい を たべる ように しています。", "english": "I make sure to eat vegetables every day."},
            {"furigana": "よる おそく に あまい もの を たべない ように します。", "english": "I will try not to eat sweet things late at night."}
        ]
    elif 'ばいいのに' in title:
        structure = "Verb [Ba-form / Conditional] + いいのに"
        nuance = "Expresses gentle advice, wish, or regret ('it would be good if...', 'you should have...')."
        samples = [
            {"furigana": "もっと はやく おきれば いい に。", "english": "You should just wake up earlier."},
            {"furigana": "いっしょ に いけば いい に。", "english": "It would be great if you could go together."}
        ]
    elif 'めく' in title:
        structure = "Noun + めく"
        nuance = "Literary N1 grammar expressing 'to carry an air of', 'to feel like', or 'to show signs of'."
        samples = [
            {"furigana": "すこしずつ はるめいて きました。", "english": "It has gradually begun to feel like spring."},
            {"furigana": "かれ の ことば に は なぞめいた 部分 が ある。", "english": "There is a mysterious air to his words."}
        ]
    else:
        # Generate furigana-only sentences for generic pattern
        samples = [
            {"furigana": f"きょう は {clean_title} について べんきょう しています。", "english": f"Today I am studying about {clean_title}."},
            {"furigana": f"かれ は {clean_title} を よく つかいます。", "english": f"He often uses the pattern {clean_title}."}
        ]

    # Conjugation attachments for grammar point backside
    conjs = {
        'root': f"【Dict Form】 + {clean_title}",
        'short': f"【Short Form】 + {clean_title}",
        'pastPos': f"【Past +ve (~た)】 + {clean_title}",
        'pastNeg': f"【Past -ve (~なかった)】 + {clean_title}",
        'teForm': f"【Te-form (~て)】 + {clean_title}",
        'taiForm': f"【Tai-form (~たい)】 + {clean_title}",
    }

    return {
        'structure': structure,
        'nuance': nuance,
        'sampleSentences': samples,
        'conjugations': conjs
    }

grammar_list = []
for sheet_name in wb.sheetnames:
    if 'Grammar' not in sheet_name:
        continue
    level = sheet_name.split()[0]
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    
    for i, r in enumerate(rows[2:]):
        if not r or not r[0]:
            continue
        title = str(r[0]).strip()
        if not title or title == 'Grammar Point':
            continue
            
        details = enrich_grammar_point(title, level)
        
        grammar_list.append({
            'id': f'g_{level.lower()}_{i}',
            'title': title,
            'jlpt': level,
            'type': 'grammar',
            'structure': details['structure'],
            'nuance': details['nuance'],
            'sampleSentences': details['sampleSentences'],
            'conjugations': details['conjugations']
        })

print(f"Extracted {len(grammar_list)} grammar entries.")

# Save outputs
os.makedirs('src/data', exist_ok=True)

with open(VOCAB_OUT, 'w', encoding='utf-8') as f:
    json.dump(vocab_list, f, ensure_ascii=False, indent=2)

with open(GRAMMAR_OUT, 'w', encoding='utf-8') as f:
    json.dump(grammar_list, f, ensure_ascii=False, indent=2)

print("Saved database files successfully!")
